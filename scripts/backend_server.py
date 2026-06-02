"""
报关资料 Excel 生成后端

依赖安装：pip install flask openpyxl

启动方式：
    python backend_server.py
"""

import base64
import json
import logging
import os
import re
import time
import zipfile
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from flask import Flask, request, jsonify, send_from_directory
from werkzeug.exceptions import HTTPException
import openpyxl
from openpyxl.cell.cell import MergedCell

app = Flask(__name__)

if not app.logger.handlers:
    logging.basicConfig(level=logging.INFO)

_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backend_access.log')
_file_handler = logging.FileHandler(_log_path, encoding='utf-8')
_file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(message)s'))
app.logger.addHandler(_file_handler)
app.logger.setLevel(logging.INFO)

_script_dir = os.path.dirname(os.path.abspath(__file__))


def _request_trace_id() -> str:
    trace_id = (request.args.get('trace_id') or request.headers.get('X-Trace-Id') or '').strip()
    return trace_id or '-'

def _first_existing(paths):
    for p in paths:
        if os.path.exists(p):
            return p
    return ''


def _pick_template(item_count: int) -> str:
    # 优先使用新模板（单商品/多商品），找不到时回退到旧模板名。
    if item_count > 1:
        candidates = [
            os.path.join(_script_dir, '报关资料模板（有多个商品）.xlsx'),
            os.path.join(_script_dir, '..', '报关资料模板（有多个商品）.xlsx'),
            os.path.join(_script_dir, '报关资料模板.xlsx'),
            os.path.join(_script_dir, '..', '报关资料模板.xlsx'),
        ]
    else:
        candidates = [
            os.path.join(_script_dir, '报关资料模板（只有一个商品）.xlsx'),
            os.path.join(_script_dir, '..', '报关资料模板（只有一个商品）.xlsx'),
            os.path.join(_script_dir, '报关资料模板.xlsx'),
            os.path.join(_script_dir, '..', '报关资料模板.xlsx'),
        ]
    template = _first_existing(candidates)
    if not template:
        raise FileNotFoundError('未找到报关资料模板文件')
    return template

OUTPUT_DIR = os.path.join(_script_dir, 'generated')
os.makedirs(OUTPUT_DIR, exist_ok=True)

REQUIRED_ROW_FIELDS = [
    '境外收货人', '合同号码', '贸易国', '运抵国', '成交方式',
    '商品编号', '品牌', '规格型号', '数量', '单位', '单价', '金额', '货源地'
]


def _apply_item_block_template_style(ws, start_row: int, row_stride: int, item_slot_count: int):
    """把首个商品块(20-22行)的样式和合并布局复制到后续商品块，避免第8项后格式漂移。"""
    style_map = {}
    for dr in range(row_stride):
        for col in range(1, 23):
            cell = ws.cell(row=start_row + dr, column=col)
            if isinstance(cell, MergedCell):
                continue
            style_map[(dr, col)] = copy(cell._style)

    row_heights = {}
    for dr in range(row_stride):
        row_heights[dr] = ws.row_dimensions[start_row + dr].height

    merge_offsets = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row >= start_row and merged.max_row <= start_row + row_stride - 1:
            merge_offsets.append((
                merged.min_row - start_row,
                merged.max_row - start_row,
                merged.min_col,
                merged.max_col,
            ))

    existing_merges = {str(m) for m in ws.merged_cells.ranges}

    for idx in range(1, item_slot_count):
        base = start_row + idx * row_stride

        for dr in range(row_stride):
            h = row_heights.get(dr)
            if h is not None:
                ws.row_dimensions[base + dr].height = h

        for dr, col in style_map:
            target = ws.cell(row=base + dr, column=col)
            if isinstance(target, MergedCell):
                continue
            target._style = copy(style_map[(dr, col)])

        for min_off, max_off, min_col, max_col in merge_offsets:
            min_row = base + min_off
            max_row = base + max_off
            ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
            if ref not in existing_merges:
                ws.merge_cells(ref)
                existing_merges.add(ref)

# HS编码 → 材质映射（来自项目计划书）
HS_MATERIAL = {
    '8507500000': '镍粉+氢氧化镍',   # 镍氢 / 镍锌
    '8506101210': '碱性锌锰',
    '8506101910': '碱性锌锰',
    '8506109010': '二氧化锰',
    '8506101110': '碱性锌锰',         # 扣式碱性
    '8506500090': '二氧化锰+铁+锂',   # 锂-二氧化锰 / 锂锰扣式 / 锂原电池
    '8506500011': '锂亚硫酰氯 ,亚硫酰氯填充小于1千克',  # ER 锂亚
    '8507600090': '锂离子',           # ICR / LP / IFR
    '8506600010': '锌粉',             # 锌空气
    '8504409999': '',                  # 充电器
    '8507200000': '',                  # 铅酸
    '8507300090': '',                  # 镍镉
}

LITHIUM_ION_HS = '8507600090'
LITHIUM_THIONYL_CHLORIDE_HS = '8506500011'

LITHIUM_THIONYL_CHLORIDE_RATIO = {
    'ER14250': Decimal('0.2710'),
    'ER10450': Decimal('0.3500'),
    'ER14335': Decimal('0.2826'),
    'ER14505': Decimal('0.3084'),
    'ER17335': Decimal('0.2683'),
    'ER17505': Decimal('0.3444'),
    'ER18505': Decimal('0.3469'),
    'ER26500': Decimal('0.3713'),
    'ER34615': Decimal('0.4040'),
    'ER341245': Decimal('0.4531'),
    'ER14250M': Decimal('0.2227'),
    'ER14335M': Decimal('0.2854'),
    'ER14505M': Decimal('0.2788'),
    'ER17335M': Decimal('0.2760'),
    'ER17505M': Decimal('0.2890'),
    'ER18505M': Decimal('0.3202'),
    'ER26500M': Decimal('0.3386'),
    'ER34615M': Decimal('0.3655'),
    'ER10280': Decimal('0.3080'),
    'ER14505H': Decimal('0.0343'),
    'ER261020': Decimal('0.3963'),
}

# HS编码 → 商品名称（短品名，填入 D 列第 0 行）
HS_PRODUCT_NAME = {
    '8507500000': '镍氢电池/镍氢电池组',
    '8506101210': '碱性电池/碱性电池组',
    '8506101910': '碱性电池',
    '8506109010': '碳性电池',
    '8506101110': '扣式电池',
    '8506500090': '锂原电池/锂原电池组',
    '8506500011': '锂电池/锂电池组',
    '8507600090': '锂电池/锂电池组',
    '8506600010': '锌空气电池',
    '8504409999': '充电器',
    '8507200000': '铅酸电池',
    '8507300090': '镍镉电池',
}

# HS编码 → 形状映射（来自项目计划书）
HS_SHAPE = {
    '8506101210': '圆柱形',   # AA/AAA 碱性圆柱
    '8506101910': '方形',
    '8506109010': '圆柱形',
    '8506101110': '纽扣形',
    '8506500090': '纽扣形',
    '8507600090': '圆柱形',
    '8506500011': '圆柱形',
}


def _parse_spec(spec: str) -> dict:
    """从规格型号字符串中提取型号、容量、电压。
    样例：PKCELL-ER26500-9000-3.6V-T-泡沫盘装
          NI-MH-AAA1000-10.8V-B-T-内盒装
    """
    parts = spec.split('-') if spec else []
    result = {'model': '', 'capacity_mah': None, 'voltage_v': None}

    if not parts:
        return result

    # 电压：找形如 3.6V / 10.8V 的片段
    for p in parts:
        m = re.match(r'^(\d+\.?\d*)V$', p)
        if m:
            result['voltage_v'] = float(m.group(1))
            break

    # 型号（位7）：第二位优先，若不满足则尝试第三位。
    # 判断：需同时包含字母和数字，兼容 4/3A3800 这类无牌型号。
    def is_model(s):
        if not s:
            return False
        if re.match(r'^\d+\.?\d*V$', s):
            return False
        return bool(re.search(r'[A-Za-z]', s) and re.search(r'\d', s))

    if len(parts) >= 2:
        if is_model(parts[1]):
            result['model'] = parts[1]
        elif len(parts) >= 3 and is_model(parts[2]):
            result['model'] = parts[2]
        else:
            result['model'] = parts[1]

    # 容量（位8）：在型号确定之后，找第三片段（相对于品牌之后）
    # 找纯数字片段 或 字母+数字末尾 的容量值
    model_idx = parts.index(result['model']) if result['model'] in parts else -1
    cap_candidates = parts[model_idx + 1:] if model_idx >= 0 else parts[2:]
    for p in cap_candidates:
        if re.match(r'^\d+$', p):             # 纯数字，如 9000
            result['capacity_mah'] = float(p)
            break
        m = re.search(r'(\d+)$', p)
        if m and not re.match(r'^\d+\.?\d*V$', p):   # 非电压片段
            result['capacity_mah'] = float(m.group(1))
            break

    # 无牌型号常见场景：容量直接包含在型号中，如 SC3000 / 4/3A3800。
    if result['capacity_mah'] is None and result['model']:
        m = re.search(r'(\d+)$', result['model'])
        if m:
            result['capacity_mah'] = float(m.group(1))

    return result


def build_product_name(row: dict, pack_qty=None, pack_net=None) -> str:
    """按项目计划书规则生成 11 段式商品名称及规格型号字符串。
    格式：位1|位2|位3|位4|位5|位6|位7|位8|位9|位10|位11
    """
    brand = str(row.get('品牌', '')).strip()
    hs_code = str(row.get('商品编号', '')).strip()
    spec = str(row.get('规格型号', '')).strip()
    is_no_brand = (not brand) or (brand in {'无牌', '无品牌', 'NOBRAND', 'NO BRAND'})

    # 位1：品牌标识
    brand_up = brand.upper()
    if 'PKCELL' in brand_up or 'PKNERGY' in brand_up:
        bit1 = '1'
    elif is_no_brand:
        bit1 = '0'
    elif brand:
        bit1 = '3'
    else:
        bit1 = '0'

    # 位2-4：固定值 / 按 HS 编码查表
    bit2 = '2'
    bit3 = '玩具用'
    bit4 = HS_SHAPE.get(hs_code, '方形')   # 默认方形

    # 位5：材质
    bit5 = HS_MATERIAL.get(hs_code, '锂离子')

    # 位6：品牌
    bit6 = brand

    # 解析规格型号
    parsed = _parse_spec(spec)
    bit7 = parsed['model']

    capacity_mah = parsed['capacity_mah']
    bit8 = f"{int(capacity_mah)}mAh" if capacity_mah is not None else ''

    bit9 = '不含汞'

    voltage_v = parsed['voltage_v']
    bit10 = f"{voltage_v}V" if voltage_v is not None else ''

    bit11 = ''
    if hs_code == LITHIUM_THIONYL_CHLORIDE_HS:
        qty_dec = _to_decimal(pack_qty if pack_qty is not None else row.get('数量', 0))
        net_dec = _to_decimal(pack_net if pack_net is not None else _first_non_empty(row, ['净重（千克）', '净重']))
        ratio = LITHIUM_THIONYL_CHLORIDE_RATIO.get(str(bit7).strip().upper())
        if qty_dec > 0 and net_dec > 0 and ratio is not None:
            content_kg = ((net_dec / qty_dec) * ratio).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
            bit11 = f'单个锂亚硫酰氯电池亚硫酰氯灌装含量{content_kg}KG'
    elif hs_code == LITHIUM_ION_HS and capacity_mah is not None and voltage_v is not None:
        qty_dec = _to_decimal(pack_qty if pack_qty is not None else row.get('数量', 0))
        net_dec = _to_decimal(pack_net if pack_net is not None else _first_non_empty(row, ['净重（千克）', '净重']))
        if qty_dec > 0 and net_dec > 0:
            wh = (Decimal(str(capacity_mah)) * Decimal(str(voltage_v))) / Decimal('1000')
            specific_energy = (wh / (net_dec / qty_dec)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            bit11 = f'比能量：{specific_energy}WH/KG'

    if bit11:
        return f'{bit1}|{bit2}|{bit3}|{bit4}|{bit5}|{bit6}|{bit7}|{bit8}|{bit9}|{bit10}|{bit11}'
    return f'{bit1}|{bit2}|{bit3}|{bit4}|{bit5}|{bit6}|{bit7}|{bit8}|{bit9}|{bit10}'


def _to_decimal(value) -> Decimal:
    """安全转换数字为 Decimal，无法解析时返回 0。"""
    if value is None:
        return Decimal('0')
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return Decimal('0')
    text = str(value).strip()
    if not text:
        return Decimal('0')
    text = text.replace(',', '')
    m = re.search(r'-?\d+(?:\.\d+)?', text)
    if not m:
        return Decimal('0')
    try:
        return Decimal(m.group(0))
    except InvalidOperation:
        return Decimal('0')


def _normalize_domestic_origin(value) -> str:
    """境内货源地标准化：深圳/珠海 -> 特区，其它地区 -> 其他。"""
    text = str(value or '').strip()
    if not text:
        return ''
    if text.endswith('特区') or text.endswith('其他'):
        return text
    base = text[:-1] if text.endswith('市') else text
    if base in {'深圳', '珠海'}:
        return base + '特区'
    return base + '其他'


def _first_non_empty(row: dict, keys) -> str:
    """从多个候选键中取第一个非空值，统一返回字符串。"""
    for key in keys:
        val = row.get(key)
        if val is None:
            continue
        text = str(val).strip()
        if text != '':
            return text
    return ''


def _copy_row_style(ws, src_row: int, dst_row: int):
    """复制整行样式（字体/边框/填充/数字格式/对齐等）及行高。"""
    src_height = ws.row_dimensions[src_row].height
    if src_height is not None:
        ws.row_dimensions[dst_row].height = src_height

    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        dst_cell._style = copy(src_cell._style)


def _copy_single_row_merges(ws, src_row: int, dst_row: int):
    """复制源行上的单行合并区到目标行。"""
    existing = {str(m) for m in ws.merged_cells.ranges}
    ranges = list(ws.merged_cells.ranges)
    for mr in ranges:
        if mr.min_row == src_row and mr.max_row == src_row:
            ref = (
                f"{openpyxl.utils.get_column_letter(mr.min_col)}{dst_row}:"
                f"{openpyxl.utils.get_column_letter(mr.max_col)}{dst_row}"
            )
            if ref not in existing:
                ws.merge_cells(ref)
                existing.add(ref)


def _safe_set_cell(ws, row: int, column: int, value):
    """仅在目标不是合并从属单元格时写值，避免 MergedCell 写入异常。"""
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _shift_merged_ranges(ws, start_row: int, delta: int):
    """将指定起始行及以下的合并区整体下移，修复 insert_rows 不移动合并区的问题。"""
    if delta <= 0:
        return
    refs = [str(mr) for mr in list(ws.merged_cells.ranges)]
    for ref in refs:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
        if min_row < start_row:
            continue
        try:
            ws.unmerge_cells(ref)
        except KeyError:
            # 某些模板在插行后会出现合并区索引不一致，忽略并继续重建。
            pass
    for ref in refs:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
        if min_row < start_row:
            continue
        new_ref = (
            f"{openpyxl.utils.get_column_letter(min_col)}{min_row + delta}:"
            f"{openpyxl.utils.get_column_letter(max_col)}{max_row + delta}"
        )
        ws.merge_cells(new_ref)


def _normalize_sheet_name(name: str) -> str:
    """标准化工作表名：去除所有空白字符，兼容模板中的异常空格。"""
    return re.sub(r'\s+', '', str(name or ''))


def _find_sheet_by_name(wb, expected_name: str):
    """按标准化名称查找工作表，找不到返回 None。"""
    target = _normalize_sheet_name(expected_name)
    for ws in wb.worksheets:
        if _normalize_sheet_name(ws.title) == target:
            return ws
    return None


def _normalize_invoice_formulas(wb, item_count: int):
    """标准化发票明细区并按商品数动态扩行，避免模板历史公式错位导致跳号。"""
    ws_invoice = _find_sheet_by_name(wb, '发票')
    if ws_invoice is None:
        return

    detail_start_row = 9
    base_capacity = 14
    summary_row = 23
    detail_count = max(1, int(item_count or 0))

    if detail_count > base_capacity:
        extra = detail_count - base_capacity
        ws_invoice.insert_rows(summary_row, extra)
        _shift_merged_ranges(ws_invoice, summary_row, extra)
        for i in range(extra):
            dst_row = summary_row + i
            _copy_row_style(ws_invoice, summary_row - 1, dst_row)
            _copy_single_row_merges(ws_invoice, summary_row - 1, dst_row)
        summary_row += extra

    last_detail_row = detail_start_row + detail_count - 1

    # 清空明细区公式位，避免旧模板残留错位引用。
    for row in range(detail_start_row, summary_row):
        for col in range(3, 9):
            ws_invoice.cell(row=row, column=col).value = None

    for item_no in range(1, detail_count + 1):
        row = detail_start_row + item_no - 1
        ws_invoice.cell(row=row, column=3).value = f'=OFFSET(报关单!$D$1,ROW(报关单!D{item_no})*3+16,0)'
        ws_invoice.cell(row=row, column=4).value = f'=OFFSET(报关单!$G$1,ROW(报关单!G{item_no})*3+18,0)'
        ws_invoice.cell(row=row, column=5).value = f'=OFFSET(报关单!$H$1,ROW(报关单!H{item_no})*3+18,0)'
        ws_invoice.cell(row=row, column=6).value = f'=IFERROR(H{row}/D{row},0)'
        ws_invoice.cell(row=row, column=7).value = f'=OFFSET(报关单!$I$1,ROW(报关单!I{item_no})*3+18,0)'
        ws_invoice.cell(row=row, column=8).value = f'=OFFSET(报关单!$I$1,ROW(报关单!I{item_no})*3+17,0)'

    # 汇总行固定重写，确保扩行后范围不偏移。
    ws_invoice.cell(row=summary_row, column=8).value = f'=SUM(H{detail_start_row}:H{last_detail_row})'


def _normalize_contract_formulas(wb, item_count: int):
    """标准化合同明细区并按商品数动态扩行，保证汇总公式范围连续。"""
    ws_contract = _find_sheet_by_name(wb, '合同')
    if ws_contract is None:
        return

    detail_start_row = 19
    base_capacity = 18
    summary_row = 37
    detail_count = max(1, int(item_count or 0))

    if detail_count > base_capacity:
        extra = detail_count - base_capacity
        ws_contract.insert_rows(summary_row, extra)
        _shift_merged_ranges(ws_contract, summary_row, extra)
        for i in range(extra):
            dst_row = summary_row + i
            _copy_row_style(ws_contract, summary_row - 1, dst_row)
            _copy_single_row_merges(ws_contract, summary_row - 1, dst_row)
        summary_row += extra

    last_detail_row = detail_start_row + detail_count - 1

    # 清空明细映射位，统一按发票明细连续映射。
    for row in range(detail_start_row, summary_row):
        ws_contract.cell(row=row, column=6).value = None
        ws_contract.cell(row=row, column=7).value = None
        ws_contract.cell(row=row, column=8).value = None

    for item_no in range(1, detail_count + 1):
        row = detail_start_row + item_no - 1
        invoice_row = 8 + item_no
        ws_contract.cell(row=row, column=6).value = f'=发票!F{invoice_row}'
        ws_contract.cell(row=row, column=7).value = f'=发票!G{invoice_row}'
        ws_contract.cell(row=row, column=8).value = f'=发票!H{invoice_row}'

    # 汇总行固定重写，兼容历史模板 I 列可能存在补充值的口径。
    ws_contract.cell(row=summary_row, column=8).value = f'=SUM(H{detail_start_row}:I{last_detail_row})'


def _is_freight_row(row: dict) -> bool:
    hs_code = str(row.get('商品编号', '')).strip()
    item_name = str(row.get('商品名称', '')).strip()
    spec_name = str(row.get('规格型号', '')).strip()
    if hs_code:
        return False
    return item_name == '国际运费' or spec_name == '国际运费'


def fill_template(rows: list) -> str:
    """填充报关单模板，返回生成文件名。rows 为同一合同号码下的所有商品行。"""
    if not rows:
        template_path = _pick_template(1)
        wb = openpyxl.load_workbook(template_path)
        filename = f'报关资料_{int(time.time())}.xlsx'
        wb.save(os.path.join(OUTPUT_DIR, filename))
        return filename

    # 识别运费行：无商品编号，且商品名称或规格型号为“国际运费”。
    freight_row = None
    item_rows = []
    for row in rows:
        if _is_freight_row(row) and freight_row is None:
            freight_row = row
            continue
        item_rows.append(row)

    template_path = _pick_template(len(item_rows) if item_rows else 1)
    wb = openpyxl.load_workbook(template_path)
    ws = wb['报关单']
    ws_entrust = _find_sheet_by_name(wb, '委托书')

    # 委托书 H16 改为动态日期公式，按生成当天自动计算。
    if ws_entrust is not None:
        ws_entrust['H16'] = '=TODAY()'

    # 买方地址后续不填写，保持留空。
    ws['V9'] = None

    first = item_rows[0] if item_rows else rows[0]

    # === 箱单字段（与报关单 E12/F12/G12 公式联动）===
    # 需求：
    # 1) 每个商品的箱数/毛重/净重按行写入箱单 C10/F10/G10 起。
    # 2) 商品超过 11 行时，自动在原汇总行前插入明细行并复制样式。
    # 3) 汇总行 C/F/G 写入 SUM 动态公式；报关单 E12/F12/G12 指向该汇总行。
    pack_summary_row = 21
    pack_start_row = 10
    ws_pack = None
    if '箱单' in wb.sheetnames:
        ws_pack = wb['箱单']
        base_capacity = 11  # C10~C20
        item_count_for_pack = len(item_rows)

        if item_count_for_pack > base_capacity:
            extra = item_count_for_pack - base_capacity
            # 在原汇总行(21)前插入新增明细行。
            ws_pack.insert_rows(21, extra)
            _shift_merged_ranges(ws_pack, 21, extra)
            # 新增行复制最后一个明细模板行(20)样式，保持格式一致。
            for i in range(extra):
                dst_row = 21 + i
                _copy_row_style(ws_pack, 20, dst_row)
                _copy_single_row_merges(ws_pack, 20, dst_row)

        # 汇总行行号：插入 extra 行后向下移动。
        pack_summary_row = 21 + max(0, item_count_for_pack - base_capacity)

        # 先清空可写入范围，避免旧值残留。
        clear_end_row = pack_summary_row - 1
        if clear_end_row >= pack_start_row:
            for r in range(pack_start_row, clear_end_row + 1):
                _safe_set_cell(ws_pack, r, 3, None)  # C: 箱数
                _safe_set_cell(ws_pack, r, 5, None)  # E: 数量单位（发票联动）
                _safe_set_cell(ws_pack, r, 6, None)  # F: 毛重
                _safe_set_cell(ws_pack, r, 7, None)  # G: 净重
                _safe_set_cell(ws_pack, r, 8, None)  # H: 核对项1
                _safe_set_cell(ws_pack, r, 9, None)  # I: 核对项2
                _safe_set_cell(ws_pack, r, 10, None) # J: 核对项3

        # 逐商品写入箱数/毛重/净重。
        for idx, row in enumerate(item_rows):
            r = pack_start_row + idx
            invoice_row = 9 + idx
            qty_text = _first_non_empty(row, ['数量'])
            carton_text = _first_non_empty(row, ['箱数', '件数'])
            gross_text = _first_non_empty(row, ['毛重（千克）', '毛重'])
            net_text = _first_non_empty(row, ['净重（千克）', '净重'])
            hs_code = str(row.get('商品编号', '')).strip()
            parsed_spec = _parse_spec(str(row.get('规格型号', '')).strip())
            cap_mah = parsed_spec.get('capacity_mah')
            voltage_v = parsed_spec.get('voltage_v')
            model_code = str(parsed_spec.get('model', '')).strip().upper()

            if qty_text != '':
                _safe_set_cell(ws_pack, r, 4, float(_to_decimal(qty_text)))
                _safe_set_cell(ws_pack, r, 5, f'=发票!E{invoice_row}')
            if carton_text != '':
                _safe_set_cell(ws_pack, r, 3, float(_to_decimal(carton_text)))
            if gross_text != '':
                _safe_set_cell(ws_pack, r, 6, float(_to_decimal(gross_text)))
            if net_text != '':
                _safe_set_cell(ws_pack, r, 7, float(_to_decimal(net_text)))

            # 核对公式：所有可计算订单先写 H=G/D；其余订单 H/I/J 保持空白。
            if hs_code in {LITHIUM_ION_HS, LITHIUM_THIONYL_CHLORIDE_HS}:
                _safe_set_cell(ws_pack, r, 8, f'=G{r}/D{r}')
                if hs_code == LITHIUM_ION_HS and cap_mah is not None and voltage_v is not None:
                    _safe_set_cell(ws_pack, r, 9, f'={float(cap_mah)}*{float(voltage_v)}/1000')
                    _safe_set_cell(ws_pack, r, 10, f'=I{r}/H{r}')
                elif hs_code == LITHIUM_THIONYL_CHLORIDE_HS:
                    ratio = LITHIUM_THIONYL_CHLORIDE_RATIO.get(model_code)
                    if ratio is not None:
                        _safe_set_cell(ws_pack, r, 9, float(ratio))
                        _safe_set_cell(ws_pack, r, 10, f'=H{r}*I{r}')

        # 汇总行 C/F/G 动态公式：=SUM(C10:C{last_row}) 等。
        last_detail_row = pack_start_row + max(item_count_for_pack, 1) - 1
        ws_pack.cell(row=pack_summary_row, column=3).value = f'=SUM(C{pack_start_row}:C{last_detail_row})'
        ws_pack.cell(row=pack_summary_row, column=6).value = f'=SUM(F{pack_start_row}:F{last_detail_row})'
        ws_pack.cell(row=pack_summary_row, column=7).value = f'=SUM(G{pack_start_row}:G{last_detail_row})'

    # === 报关单表头字段 ===
    contract_no = str(first.get('合同号码', '')).strip()
    if contract_no.upper().startswith('BK'):
        ws['A4'] = '深圳市倍苛新能源有限公司'
    ws['A6']  = first.get('境外收货人', '')    # 境外收货人（D14 是公式 =A6，不覆盖）
    ws['A10'] = contract_no                     # 合同协议号
    ws['E10'] = first.get('贸易国', '')         # 贸易国（地区）
    ws['G10'] = first.get('运抵国', '') or first.get('贸易国', '')  # 运抵国（地区）
    ws['I12'] = first.get('成交方式', '')       # 成交方式在 I12（I11 是标签行）

    # 运费写入：值写入 K12（K12:L12 为合并单元格，写主单元格 K12）。
    freight_amount = Decimal('0')
    if freight_row is not None:
        freight_price = str(freight_row.get('单价', '')).strip()
        freight_currency = str(freight_row.get('币制') or freight_row.get('币别') or '').strip()
        freight_text = (freight_price + (' ' + freight_currency if freight_currency else '')).strip()
        freight_amount = _to_decimal(freight_row.get('单价', 0))
        if freight_text:
            ws['K12'] = freight_text

    # === 商品明细：每行一条商品，从第20行起，每条占3行 ===
    # 若模板实际行距不同，请调整 ROW_STRIDE
    ROW_STRIDE = 3
    START_ROW = 20   # 1-based
    ITEM_SLOT_COUNT = 80

    # 第8项后的格式（字体、边框、合并区）也要与前面一致。
    _apply_item_block_template_style(ws, START_ROW, ROW_STRIDE, ITEM_SLOT_COUNT)

    # 多商品模板可能自带示例数据。先清理可写入位，避免残留“多余项号”。
    # 只清理程序会写入的单元格，不动公式位（如 I(base+1)）。
    for idx in range(ITEM_SLOT_COUNT):
        base = START_ROW + idx * ROW_STRIDE
        ws.cell(row=base, column=1).value = None       # A: 项号
        ws.cell(row=base, column=2).value = None       # B: 商品编号
        ws.cell(row=base + 1, column=2).value = None   # B+1: 征免（固定 102）
        ws.cell(row=base, column=4).value = None       # D: 商品名称
        ws.cell(row=base + 1, column=4).value = None   # D+1: 规格型号
        ws.cell(row=base + 2, column=7).value = None   # G+2: 数量
        ws.cell(row=base + 2, column=8).value = None   # H+2: 单位
        ws.cell(row=base, column=9).value = None       # I: 单价
        ws.cell(row=base + 1, column=9).value = None   # I+1: 总价（公式）
        ws.cell(row=base + 2, column=9).value = None   # I+2: 币制
        ws.cell(row=base, column=11).value = None      # K: 原产国
        ws.cell(row=base, column=13).value = None      # M: 最终目的国（地区）
        ws.cell(row=base, column=16).value = None      # P: 境内货源地
        ws.cell(row=base, column=19).value = None      # S: 征税方式

    # 运费分摊规则：有运费时按数量占比分摊到每个商品总价，再反推单价。
    item_qty_list = []
    for row in item_rows:
        item_qty_list.append(_to_decimal(row.get('数量', 0)))
    total_qty = sum(item_qty_list, Decimal('0'))

    allocated_freight_list = [Decimal('0')] * len(item_rows)
    if freight_amount > 0 and total_qty > 0 and item_rows:
        allocated_sum = Decimal('0')
        for idx in range(len(item_rows)):
            if idx == len(item_rows) - 1:
                allocated = (freight_amount - allocated_sum).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            else:
                allocated = (freight_amount * item_qty_list[idx] / total_qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                allocated_sum += allocated
            allocated_freight_list[idx] = allocated

    for idx, row in enumerate(item_rows):
        base = START_ROW + idx * ROW_STRIDE   # openpyxl 行号从 1 开始
        hs_code = str(row.get('商品编号', '')).strip()
        short_name = str(row.get('商品名称', '')).strip() or HS_PRODUCT_NAME.get(hs_code, hs_code)

        pack_qty_value = None
        pack_net_value = None
        if ws_pack is not None:
            pack_row = pack_start_row + idx
            pack_qty_value = ws_pack.cell(row=pack_row, column=4).value
            pack_net_value = ws_pack.cell(row=pack_row, column=7).value

        spec_string = build_product_name(row, pack_qty=pack_qty_value, pack_net=pack_net_value)  # 规格型号（11段）

        qty_dec = item_qty_list[idx]
        qty = float(qty_dec)

        unit_price_dec = _to_decimal(row.get('单价', 0))
        base_total_dec = (qty_dec * unit_price_dec).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        allocated_freight = allocated_freight_list[idx]

        if freight_amount > 0 and total_qty > 0 and qty_dec > 0:
            final_total_dec = (base_total_dec + allocated_freight).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            final_unit_dec = (final_total_dec / qty_dec).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
            total_cell_value = float(final_total_dec)
        else:
            final_unit_dec = unit_price_dec.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
            total_cell_value = f'=G{base + 2}*I{base}'

        ws.cell(row=base,     column=1).value  = idx + 1      # A: 项号
        ws.cell(row=base,     column=2).value  = hs_code      # B: 商品编号（HS）
        ws.cell(row=base + 1, column=2).value  = '102'        # B+1: 征免方式
        ws.cell(row=base,     column=4).value  = short_name   # D: 商品名称（短品名）
        ws.cell(row=base + 1, column=4).value  = spec_string  # D (合并行+1): 规格型号（11段）
        ws.cell(row=base + 2, column=7).value  = qty                      # G: 数量
        ws.cell(row=base + 2, column=8).value  = row.get('单位', '')       # H: 单位
        ws.cell(row=base,     column=9).value  = float(final_unit_dec)     # I: 单价（有运费时为分摊后单价，保留4位）
        ws.cell(row=base + 1, column=9).value  = total_cell_value           # I+1: 总价（有运费时为分摊后总价，保留2位）
        ws.cell(row=base + 2, column=9).value  = row.get('币制', '美元')   # I: 币制
        ws.cell(row=base,     column=11).value = '中国'                    # K: 原产国
        ws.cell(row=base,     column=13).value = row.get('运抵国', '') or first.get('运抵国', '')  # M: 最终目的国（地区）
        ws.cell(row=base,     column=16).value = _normalize_domestic_origin(row.get('货源地', ''))  # P: 境内货源地
        ws.cell(row=base,     column=19).value = '照章征税'                # S: 征税方式

        # 商品明细中的毛重由人工填写：清空 U 列值位。
        ws.cell(row=base, column=21).value = None

    # 按实际商品数删除多余明细行，避免模板中未使用的占位区域出现在导出结果中。
    used_count = len(item_rows)
    if used_count > 0 and used_count < ITEM_SLOT_COUNT:
        delete_start = START_ROW + used_count * ROW_STRIDE
        delete_count = (ITEM_SLOT_COUNT - used_count) * ROW_STRIDE
        ws.delete_rows(delete_start, delete_count)

    # 报关单件数/毛重/净重公式联动到箱单汇总行（汇总行会随商品数动态调整）。
    ws['E12'] = f'=箱单!C{pack_summary_row}'  # 件数
    ws['F12'] = f'=箱单!F{pack_summary_row}'  # 毛重（千克）
    ws['G12'] = f'=箱单!G{pack_summary_row}'  # 净重（千克）

    # 统一重写发票/合同明细区公式并按商品数扩行，修复模板中历史跳号。
    _normalize_invoice_formulas(wb, len(item_rows))
    _normalize_contract_formulas(wb, len(item_rows))

    filename = f'报关资料_{int(time.time())}.xlsx'
    wb.save(os.path.join(OUTPUT_DIR, filename))
    return filename


def _validate_rows(rows: list):
    if not isinstance(rows, list) or not rows:
        return False, 'rows 为空或格式不正确'
    if not any(isinstance(item, dict) for item in rows):
        return False, 'rows 中无有效对象'
    # 非阻断策略：字段缺失/值为空不拦截，由模板留空并通过备注提示人工补齐。
    return True, ''


def _normalize_rows(rows: list):
    normalized = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        item = dict(row)
        if not item.get('币制') and item.get('币别'):
            item['币制'] = item.get('币别')
        normalized.append(item)
    return normalized



def _add_cors(response):
    # BunnyCDN 已自动注入 Allow-Origin 和 Allow-Headers，这里只补 Allow-Methods（BunnyCDN 不加这个）
    response.headers['Access-Control-Allow-Methods'] = 'POST, GET, OPTIONS'
    return response


@app.before_request
def log_request_start():
    request._start_ts = time.time()
    d = request.args.get('d', '')
    trace_id = _request_trace_id()
    app.logger.info(
        'REQ trace_id=%s method=%s path=%s remote=%s host=%s ua=%s d_len=%s',
        trace_id,
        request.method,
        request.path,
        request.headers.get('X-Forwarded-For', request.remote_addr),
        request.host,
        request.headers.get('User-Agent', '')[:180],
        len(d) if d else 0,
    )


@app.after_request
def after_request(response):
    took_ms = int((time.time() - getattr(request, '_start_ts', time.time())) * 1000)
    app.logger.info(
        'RES trace_id=%s method=%s path=%s status=%s took_ms=%s',
        _request_trace_id(),
        request.method,
        request.path,
        response.status_code,
        took_ms,
    )
    return _add_cors(response)


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        return exc
    trace_id = _request_trace_id()
    app.logger.exception('UNHANDLED trace_id=%s method=%s path=%s code=BG5000 err=%s', trace_id, request.method, request.path, str(exc))
    return f'BG5000 服务器内部错误（traceId={trace_id}）', 500


@app.route('/generate', methods=['GET', 'POST', 'OPTIONS'])
def generate():
    if request.method == 'OPTIONS':
        return _add_cors(app.response_class(status=204))

    if request.method == 'GET':
        # GET ?d=<url-safe-base64-json>  →  直接下载 Excel
        # 钉钉脚本无法做异步 HTTP，改为把数据编码进 URL，用户点击链接时服务器解码生成文件
        d = request.args.get('d', '').strip()
        trace_id = _request_trace_id()
        if not d:
            app.logger.warning('ERR trace_id=%s code=BG4001 msg=missing_d', trace_id)
            return f'BG4001 缺少参数 d（traceId={trace_id}）', 400
        try:
            padded = d.replace('-', '+').replace('_', '/')
            padded += '=' * (-len(padded) % 4)
            data_bytes = base64.b64decode(padded)
            data = json.loads(data_bytes.decode('utf-8'))
            rows = data.get('rows', [])
        except Exception as exc:
            app.logger.warning('ERR trace_id=%s code=BG4002 msg=decode_failed detail=%s', trace_id, str(exc))
            return f'BG4002 数据解码失败（traceId={trace_id}）: {exc}', 400
        ok, msg = _validate_rows(rows)
        if not ok:
            app.logger.warning('ERR trace_id=%s code=BG4003 msg=validate_failed detail=%s', trace_id, msg)
            return f'BG4003 {msg}（traceId={trace_id}）', 400
        rows = _normalize_rows(rows)
        contract_no = ''
        if rows and isinstance(rows[0], dict):
            contract_no = str(rows[0].get('合同号码', '')).strip()
        app.logger.info('BIZ trace_id=%s event=generate_start contract_no=%s row_count=%s', trace_id, contract_no, len(rows))
        filename = fill_template(rows)
        app.logger.info('BIZ trace_id=%s event=generate_done contract_no=%s filename=%s', trace_id, contract_no, filename)
        return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)

    # POST: 原有逻辑（保留兼容性）
    data = request.get_json(force=True)
    trace_id = _request_trace_id()
    rows = data.get('rows', [])
    ok, msg = _validate_rows(rows)
    if not ok:
        app.logger.warning('ERR trace_id=%s code=BG4003 msg=validate_failed detail=%s', trace_id, msg)
        return jsonify({'error': msg, 'code': 'BG4003', 'traceId': trace_id}), 400
    rows = _normalize_rows(rows)
    public_base = os.environ.get('PUBLIC_BASE_URL', request.host_url.rstrip('/'))
    contract_no = ''
    if rows and isinstance(rows[0], dict):
        contract_no = str(rows[0].get('合同号码', '')).strip()
    app.logger.info('BIZ trace_id=%s event=generate_start contract_no=%s row_count=%s', trace_id, contract_no, len(rows))
    filename = fill_template(rows)
    app.logger.info('BIZ trace_id=%s event=generate_done contract_no=%s filename=%s', trace_id, contract_no, filename)
    url = f'{public_base}/download/{filename}'
    return jsonify({'url': url, 'filename': filename, 'traceId': trace_id})



@app.route('/download/<filename>')
def download(filename):
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


@app.route('/gen')
def generate_from_link():
    """GET /gen?d=<url-safe-base64-json>  →  直接下载 Excel。
    钉钉脚本无法做异步 HTTP，改为把数据编码进 URL，用户点击链接时服务器解码生成文件。
    """
    d = request.args.get('d', '').strip()
    if not d:
        return '缺少参数 d', 400
    try:
        # URL-safe base64 → 标准 base64，补齐 padding
        padded = d.replace('-', '+').replace('_', '/')
        padded += '=' * (-len(padded) % 4)
        data_bytes = base64.b64decode(padded)
        data = json.loads(data_bytes.decode('utf-8'))
        rows = data.get('rows', [])
    except Exception as exc:
        return f'数据解码失败: {exc}', 400

    ok, msg = _validate_rows(rows)
    if not ok:
        return msg, 400
    rows = _normalize_rows(rows)

    filename = fill_template(rows)
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


if __name__ == '__main__':
    print('后端已启动，监听 http://0.0.0.0:5000')
    print('生成的文件保存在:', os.path.abspath(OUTPUT_DIR))
    app.run(host='0.0.0.0', port=5000, debug=False)
